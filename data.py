import openpyxl


class Airport:
    def __init__(self, airport_id: str, rating: float, name: str):
        self.airport_id = airport_id
        self.rating = rating
        self.name = name


class Airline:
    def __init__(self, airline_id: str, name: str, rating: int):
        self.airline_id = airline_id
        self.name = name
        self.rating = rating


class Route:
    def __init__(self, airline_id: str, destination_id: str, source_id: str):
        self.airline_id = airline_id
        self.destination_id = destination_id
        self.source_id = source_id


def create_list_of_airlines() -> list:
    wb = openpyxl.load_workbook(r'C:\Users\jasie\STUDIA\TASS\projekt2\TASS\airlines_ratings.xlsx')
    sheet = wb.active
    l = []
    sheet_length = len(sheet['A'])
    for i in range(2, sheet_length + 1):
        name = sheet["B" + str(i)]
        rating = sheet["G" + str(i)]
        airline_id = sheet["A" + str(i)]
        airline = Airline(airline_id=airline_id.value, name=name.value, rating=rating.value)
        l.append(airline)
    return l


def get_airline_rating_by_id(airlines: list, airline_id: str) -> float:
    for airline in airlines:
        if str(airline.airline_id) == airline_id:
            return airline.rating
    return -1.0


def get_airline_name_by_id(airlines: list, airline_id: str) -> str:
    for airline in airlines:
        if str(airline.airline_id) == airline_id:
            return airline.name
    return ""


def create_list_of_airports(routes: list[Route]) -> list:
    wb = openpyxl.load_workbook(r'C:\Users\jasie\STUDIA\TASS\projekt2\TASS\airports_ratings.xlsx')
    sheet = wb.active
    l = []
    sheet_length = len(sheet['A'])
    for i in range(2, sheet_length + 1):
        name = sheet["B" + str(i)]
        rating = sheet["J" + str(i)]
        airport_id = sheet["A" + str(i)]
        if check_if_airport_is_on_any_route(routes=routes, airport_id=airport_id.value):
            airport = Airport(airport_id=airport_id.value, name=name.value, rating=rating.value)
            l.append(airport)
    return l


def get_airport_rating_by_id(airports: list, airport_id: str) -> float:
    for airport in airports:
        if str(airport.airport_id) == airport_id:
            return airport.rating
    return -1.0


def get_airport_name_by_id(airports: list, airport_id: str) -> str:
    for airport in airports:
        if str(airport.airport_id) == airport_id:
            return airport.name
    return ""


def create_list_of_routes() -> list:
    wb = openpyxl.load_workbook(r'C:\Users\jasie\STUDIA\TASS\projekt2\TASS\routes1.xlsx')
    sheet = wb.active
    l = []
    sheet_length = len(sheet['A'])
    for i in range(2, sheet_length + 1):
        airline_id = sheet["B" + str(i)]
        destination = sheet["F" + str(i)]
        source = sheet["D" + str(i)]
        route = Route(airline_id=airline_id.value, destination_id=destination.value, source_id=source.value)
        l.append(route)
    return l


def check_if_airport_is_on_any_route(routes: list[Route], airport_id: str) -> bool:
    for route in routes:
        if route.destination_id == airport_id:
            return True
        elif route.source_id == airport_id:
            return True
    return False


def get_airlines_by_destination_and_source(routes: list, destination: str, source: str) -> list:
    l = set()
    for route in routes:
        if ((str(route.source_id) == source) and (str(route.destination_id) == destination)) or (
                (str(route.source_id) == destination) and (str(route.destination_id) == source)):
            l.add(route.airline_id)
    return l


def get_id_by_airport_name(airports: list, airport_name: str) -> str:
    for airport in airports:
        if str(airport.name) == airport_name:
            return airport.airport_id
    return -1.0
